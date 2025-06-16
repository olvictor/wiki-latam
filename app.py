from collections import defaultdict
import os
from flask import Flask, request, send_from_directory
import pandas as pd
import re
import unicodedata
from flask_cors import CORS
from flask import Flask, render_template, jsonify, request, redirect, flash, get_flashed_messages
from openpyxl import load_workbook
from apscheduler.schedulers.background import BackgroundScheduler
from flask_sqlalchemy import SQLAlchemy
import os
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from datetime import datetime
import urllib.request
import urllib.parse 
from urllib.parse import urlparse
import json


app = Flask(__name__)
CORS(app)


app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.secret_key= "#aSDa32#"
db = SQLAlchemy(app)

class Link(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    url = db.Column(db.String(500), nullable=False)
    aprovado = db.Column(db.Boolean, default=False)
    data = db.Column(db.DateTime, default=datetime.utcnow)



CLIENT_ID = 'qdn4nkcag974fxftx5xwfs5goddqx6'
CLIENT_SECRET = 'l0a2yo19o38jl7luid7nu0xerz38m1'
YOUTUBE_API_KEY = 'AIzaSyBjK1M3gsWEZCA75qYeIXBFzgoVu7vKcqY'
API_SERVICE_NAME = 'youtube'
API_VERSION = 'v3'

stream_cache = []

CHANNEL_IDS = [
    {"dados": "Cabana do Sentinela", "links": "https://www.twitch.tv/cabanadosentinela","imagem": "assets/classes/sentinela.png","imagem_gif": "assets/classes/andando/sentinela.gif","status": "off","imagem_sentado": "assets/classes/sentados/sentinela.gif","plataforma" :"youtube","channel_id":"UCyrRWvII61mthxZlbK3OFJQ"},
    {"dados": "Estudo arcano", "links": "https://www.youtube.com/@estudoarcano","imagem": "assets/classes/arcano.png","imagem_gif": "assets/classes/andando/arcano.gif","status": "off","imagem_sentado": "assets/classes/sentados/arcano.gif","plataforma" :"youtube","channel_id":"UCyCkaDZmkJdOcfwtQSYJ5OQ"},
    {"dados": "Jeff da Gaita", "links": "https://www.youtube.com/@JeffodaGaita","imagem": "assets/classes/renegado.png","imagem_gif": "assets/classes/andando/renegado.gif","status": "off","imagem_sentado": "assets/classes/sentados/renegado.gif","plataforma" :"youtube","channel_id":"UC4B7uZNcTQG2iJkSrYVztQg"},
    {"dados": "Joga Junto Ragnarok", "links": "https://www.youtube.com/@JogaJuntoRagnarok","imagem": "assets/classes/cavaleiro_runico.png","imagem_gif": "assets/classes/andando/cavaleiro_runico.gif","status": "off","imagem_sentado": "assets/classes/sentados/cavaleiro_runico.gif","plataforma" :"youtube","channel_id":"UCmLixl7G_IxDo6nl4vYCHjA"},
]


def get_youtube_service():
    return build(API_SERVICE_NAME, API_VERSION, developerKey=YOUTUBE_API_KEY)


def check_channel_live_status(youtube_service, channel_id):
    """Verifica se um canal do YouTube está transmitindo ao vivo e retorna dados da live."""
    print(f"\nVerificando o canal com ID: {channel_id}...")
    try:
        search_response = youtube_service.search().list(
            channelId=channel_id,
            eventType='live',
            type='video',
            part='id,snippet',
            maxResults=1
        ).execute()

        if search_response.get('items'):
            item = search_response['items'][0]
            video_id = item['id']['videoId']
            title = item['snippet']['title']
            thumbnail_url = item['snippet']['thumbnails']['high']['url']
            category = item['snippet'].get('categoryId', 'Live') 

            video_response = youtube_service.videos().list(
                part="liveStreamingDetails,statistics",
                id=video_id
            ).execute()

            video_info = video_response['items'][0]
            viewer_count = video_info.get('liveStreamingDetails', {}).get('concurrentViewers', 'N/A')
            started_at = video_info.get('liveStreamingDetails', {}).get('actualStartTime')

            return True, {
                'title': title,
                'category': category,
                'viewer_count': viewer_count,
                'started_at': started_at,
                'thumbnail_url': thumbnail_url
            }

        else:
            print(f"❌ O canal {channel_id} não está ao vivo.")
            return False, {}

    except HttpError as e:
        print(f"Erro HTTP ao acessar API para o canal {channel_id}: {e}")
        return False, {}
    except Exception as e:
        print(f"Erro inesperado para o canal {channel_id}: {e}")
        return False, {}


def get_access_token():
    url = 'https://id.twitch.tv/oauth2/token'
    data = urllib.parse.urlencode({
        'client_id': CLIENT_ID,
        'client_secret': CLIENT_SECRET,
        'grant_type': 'client_credentials'
    }).encode('utf-8')

    req = urllib.request.Request(url, data=data, method='POST')
    with urllib.request.urlopen(req) as response:
        result = json.loads(response.read().decode())
        return result['access_token']
       
def get_stream_data(username):
    token = get_access_token()

    url = f'https://api.twitch.tv/helix/streams?user_login={username}'
    req = urllib.request.Request(url)
    req.add_header('Client-ID', CLIENT_ID)
    req.add_header('Authorization', f'Bearer {token}')

    try:
        with urllib.request.urlopen(req) as response:
            result = json.loads(response.read().decode())
            streams = result.get('data', [])

            if streams:
                return streams[0]
            else:
                return None
    except urllib.error.HTTPError as e:
        print(f"Erro HTTP {e.code}: {e.reason}")
        print(f"Resposta: {e.read()}")
        return None

def extrair_nome_usuario(url):
    match = re.search(r'https://www.twitch.tv/([a-zA-Z0-9_]+)', url)
    if match:
        return match.group(1)
    return None  

def atualizar_stream_cache():
    global stream_cache
    print("Atualizando cache de streamers...")

    file_path = os.path.join(os.path.dirname(__file__), 'rola.xlsx')

    try:
        df = pd.read_excel(file_path, sheet_name='INFORMAÇÕES', header=None)
    except Exception as e:
        print(f"Erro ao ler o Excel: {e}")
        return

    dados = df.iloc[19:, 1].dropna().tolist()

    wb = load_workbook(file_path)
    ws = wb["INFORMAÇÕES"]
    
    links = []
    for row in ws.iter_rows(min_row=20, max_row=31, min_col=2, max_col=2):
        cell = row[0]
        if cell.hyperlink:
            links.append(cell.hyperlink.target)
        elif cell.value:
            links.append(cell.value)

    dados_links_imagens = [
         {"dados": dados[0].split(':')[0].strip(), "links": links[0],"imagem": "assets/classes/feiticeiro.png","imagem_gif": "assets/classes/andando/feiticeiro.gif","status": "off","imagem_sentado": "assets/classes/sentados/feiticeiro.gif","plataforma" :"twitch"},
         {"dados": dados[1].split(':')[0].strip(), "links": links[1],"imagem": "assets/classes/sura.png","imagem_gif": "assets/classes/andando/sura.gif","status": "off","imagem_sentado": "assets/classes/sentados/shura.gif","plataforma" :"twitch"},
        #  {"dados": dados[2].split(':')[0].strip(), "links": links[2],"imagem": "assets/classes/arcano.png","imagem_gif": "assets/classes/gifs/cavaleiro_runico.gif","status": "off"},
        #  {"dados": dados[3].split(':')[0].strip(), "links": links[3],"imagem": "assets/classes/sicario.png","imagem_gif": "assets/classes/gifs/cavaleiro_runico.gif","status": "off"},   
         {"dados": dados[4].split(':')[0].strip(), "links": links[4],"imagem": "assets/classes/arcebispo.png","imagem_gif": "assets/classes/andando/arcebispo.gif","status": "off","imagem_sentado": "assets/classes/sentados/arcebispo.gif","plataforma" :"twitch"},   
         {"dados": dados[5].split(':')[0].strip(), "links": links[5],"imagem": "assets/classes/trovador.png","imagem_gif": "assets/classes/andando/trovador.gif","status": "off","imagem_sentado": "assets/classes/sentados/trovador.gif","plataforma" :"twitch"},    # lyelz
         {"dados": dados[6].split(':')[0].strip(), "links": links[6],"imagem": "assets/classes/guardioes_reais.png","imagem_gif": "assets/classes/andando/guardiao_real.gif","status": "off","imagem_sentado": "assets/classes/sentados/guardiao_real.gif","plataforma" :"twitch"}, 
         {"dados": dados[7].split(':')[0].strip(), "links": links[7],"imagem": "assets/classes/cavaleiro_runico.png","imagem_gif": "assets/classes/andando/cavaleiro_runico.gif","status": "off","imagem_sentado": "assets/classes/sentados/cavaleiro_runico.gif","plataforma" :"twitch"},  #Asbrun 
         {"dados": dados[8].split(':')[0].strip(), "links": links[8],"imagem": "assets/classes/renegado.png","imagem_gif": "assets/classes/andando/cavaleiro_runico.gif","status": "off","imagem_sentado": "assets/classes/sentados/renegado.gif","plataforma" :"twitch"},   
         {"dados": dados[9].split(':')[0].strip(), "links": links[9],"imagem": "assets/classes/arcano.png","imagem_gif": "assets/classes/andando/arcano.gif","status": "off","imagem_sentado": "assets/classes/sentados/arcano.gif","plataforma" :"twitch"},
         {"dados": dados[10].split(':')[0].strip(), "links": links[10],"imagem": "assets/classes/sentinela.png","imagem_gif": "assets/classes/andando/trovador.gif","status": "off","imagem_sentado": "assets/classes/sentados/sentinela.gif","plataforma" :"twitch"}, 
         {"dados": "Alvaro TV", "links": "https://www.twitch.tv/alvarotv23","imagem": "assets/classes/sicario.png","imagem_gif": "assets/classes/andando/sicario.gif","status": "off","imagem_sentado": "assets/classes/sentados/sicario.gif","plataforma" :"twitch"},   
         {"dados": "Slash", "links": "https://www.twitch.tv/slashvidal","imagem": "assets/classes/sicario.png","imagem_gif": "assets/classes/andando/sicario.gif","status": "off","imagem_sentado": "assets/classes/sentados/sicario.gif","plataforma" :"twitch"},   
          {"dados": "Cabana do Sentinela", "links": "https://www.twitch.tv/cabanadosentinela","imagem": "assets/classes/sentinela.png","imagem_gif": "assets/classes/andando/sentinela.gif","status": "off","imagem_sentado": "assets/classes/sentados/sentinela.gif","plataforma" :"youtube","channel_id":"UCyrRWvII61mthxZlbK3OFJQ"},
    {"dados": "Estudo arcano", "links": "https://www.youtube.com/@estudoarcano","imagem": "assets/classes/arcano.png","imagem_gif": "assets/classes/andando/arcano.gif","status": "off","imagem_sentado": "assets/classes/sentados/arcano.gif","plataforma" :"youtube","channel_id":"UCyCkaDZmkJdOcfwtQSYJ5OQ"},
    {"dados": "Jeff da Gaita", "links": "https://www.youtube.com/@JeffodaGaita","imagem": "assets/classes/renegado.png","imagem_gif": "assets/classes/andando/renegado.gif","status": "off","imagem_sentado": "assets/classes/sentados/renegado.gif","plataforma" :"youtube","channel_id":"UC4B7uZNcTQG2iJkSrYVztQg"},
    {"dados": "Joga Junto Ragnarok", "links": "https://www.youtube.com/@JogaJuntoRagnarok","imagem": "assets/classes/cavaleiro_runico.png","imagem_gif": "assets/classes/andando/cavaleiro_runico.gif","status": "off","imagem_sentado": "assets/classes/sentados/cavaleiro_runico.gif","plataforma" :"youtube","channel_id":"UCmLixl7G_IxDo6nl4vYCHjA"},
    {"dados": "Asenhorita_any", "links": "https://www.twitch.tv/asenhorita_any","imagem": "assets/classes/musa.png","imagem_gif": "assets/classes/andando/musa.gif","status": "off","imagem_sentado": "assets/classes/sentados/musa.gif","plataforma" :"twitch"},


    ]
    youtube_service = get_youtube_service()
    for item in dados_links_imagens:
        plataforma = item.get('plataforma')

        if plataforma == 'twitch':
            username = extrair_nome_usuario(item['links'])

            if username:
                try:
                    stream_data = get_stream_data(username)
                    if stream_data:
                        item['status'] = 'on'
                        item['stream_info'] = {
                            'title': stream_data.get('title'),
                            'game_name': stream_data.get('game_name'),
                            'viewer_count': stream_data.get('viewer_count'),
                            'started_at': stream_data.get('started_at'),
                            'language': stream_data.get('language'),
                            'thumbnail_url': stream_data.get('thumbnail_url')
                        }
                    else:
                        item['status'] = 'off'
                except Exception as e:
                    print(f"Erro ao verificar status de {username} na Twitch: {e}")
                    item['status'] = 'off'
            else:
                print(f"Nome de usuário Twitch não encontrado na URL: {item['links']}")

        elif plataforma == 'youtube':
            channel_id = item['channel_id']

            if channel_id:
                try:
                    is_live, video_data = check_channel_live_status(youtube_service, channel_id)  
                    if is_live:
                        item['status'] = 'on'
                        item['stream_info'] = {
                            'title': video_data.get('title'),
                            'game_name': video_data.get('category', 'Live'),  
                            'viewer_count': video_data.get('viewer_count', 'N/A'),
                            'started_at': video_data.get('started_at'),
                            'thumbnail_url': video_data.get('thumbnail_url')
                        }
                    else:
                        item['status'] = 'off'
                except Exception as e:
                    print(f"Erro ao verificar status de {channel_id} no YouTube: {e}")
                    item['status'] = 'off'
            else:
                print(f"Channel ID do YouTube não encontrado na URL: {item['links']}")
    stream_cache =dados_links_imagens
scheduler = BackgroundScheduler()
scheduler.add_job(func=atualizar_stream_cache, trigger="interval", seconds=1800)
scheduler.start()

atualizar_stream_cache() 

def check_live(username):
        try:
            live, data = is_stream_live(username)
            return jsonify({
                'username': username,
                'live': live,
                'stream_data': data
            })
        except Exception as e:
            return jsonify({'error': str(e)}), 500

def carregar_links():
    file_path = 'rola.xlsx'
    df = pd.read_excel(file_path, header=None)

    links = df.iloc[10:14, 1].dropna().tolist()

    icones = {
        'Site': '''
            <svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke-width="1.5" stroke="currentColor" class="size-6">
                <path stroke-linecap="round" stroke-linejoin="round" d="m6.115 5.19.319 1.913A6 6 0 0 0 8.11 10.36L9.75 12l-.387.775c-.217.433-.132.956.21 1.298l1.348 1.348c.21.21.329.497.329.795v1.089c0 .426.24.815.622 1.006l.153.076c.433.217.956.132 1.298-.21l.723-.723a8.7 8.7 0 0 0 2.288-4.042 1.087 1.087 0 0 0-.358-1.099l-1.33-1.108c-.251-.21-.582-.299-.905-.245l-1.17.195a1.125 1.125 0 0 1-.98-.314l-.295-.295a1.125 1.125 0 0 1 0-1.591l.13-.132a1.125 1.125 0 0 1 1.3-.21l.603.302a.809.809 0 0 0 1.086-1.086L14.25 7.5l1.256-.837a4.5 4.5 0 0 0 1.528-1.732l.146-.292M6.115 5.19A9 9 0 1 0 17.18 4.64M6.115 5.19A8.965 8.965 0 0 1 12 3c1.929 0 3.716.607 5.18 1.64" />
            </svg>

        ''',
        'Discord': '''
           <svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke-width="1.5" stroke="currentColor" class="size-6">
                <path stroke-linecap="round" stroke-linejoin="round" d="M7.5 8.25h9m-9 3H12m-9.75 1.51c0 1.6 1.123 2.994 2.707 3.227 1.129.166 2.27.293 3.423.379.35.026.67.21.865.501L12 21l2.755-4.133a1.14 1.14 0 0 1 .865-.501 48.172 48.172 0 0 0 3.423-.379c1.584-.233 2.707-1.626 2.707-3.228V6.741c0-1.602-1.123-2.995-2.707-3.228A48.394 48.394 0 0 0 12 3c-2.392 0-4.744.175-7.043.513C3.373 3.746 2.25 5.14 2.25 6.741v6.018Z" />
           </svg>

        ''',
        'Calculadora': '''
           <svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke-width="1.5" stroke="currentColor" class="size-6">
                <path stroke-linecap="round" stroke-linejoin="round" d="M15.75 15.75V18m-7.5-6.75h.008v.008H8.25v-.008Zm0 2.25h.008v.008H8.25V13.5Zm0 2.25h.008v.008H8.25v-.008Zm0 2.25h.008v.008H8.25V18Zm2.498-6.75h.007v.008h-.007v-.008Zm0 2.25h.007v.008h-.007V13.5Zm0 2.25h.007v.008h-.007v-.008Zm0 2.25h.007v.008h-.007V18Zm2.504-6.75h.008v.008h-.008v-.008Zm0 2.25h.008v.008h-.008V13.5Zm0 2.25h.008v.008h-.008v-.008Zm0 2.25h.008v.008h-.008V18Zm2.498-6.75h.008v.008h-.008v-.008Zm0 2.25h.008v.008h-.008V13.5ZM8.25 6h7.5v2.25h-7.5V6ZM12 2.25c-1.892 0-3.758.11-5.593.322C5.307 2.7 4.5 3.65 4.5 4.757V19.5a2.25 2.25 0 0 0 2.25 2.25h10.5a2.25 2.25 0 0 0 2.25-2.25V4.757c0-1.108-.806-2.057-1.907-2.185A48.507 48.507 0 0 0 12 2.25Z" />
           </svg>

        ''',
        'Skill Simulator': '''
            <svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke-width="1.5" stroke="currentColor" class="size-6">
                <path stroke-linecap="round" stroke-linejoin="round" d="M12 3v17.25m0 0c-1.472 0-2.882.265-4.185.75M12 20.25c1.472 0 2.882.265 4.185.75M18.75 4.97A48.416 48.416 0 0 0 12 4.5c-2.291 0-4.545.16-6.75.47m13.5 0c1.01.143 2.01.317 3 .52m-3-.52 2.62 10.726c.122.499-.106 1.028-.589 1.202a5.988 5.988 0 0 1-2.031.352 5.988 5.988 0 0 1-2.031-.352c-.483-.174-.711-.703-.59-1.202L18.75 4.971Zm-16.5.52c.99-.203 1.99-.377 3-.52m0 0 2.62 10.726c.122.499-.106 1.028-.589 1.202a5.989 5.989 0 0 1-2.031.352 5.989 5.989 0 0 1-2.031-.352c-.483-.174-.711-.703-.59-1.202L5.25 4.971Z" />
            </svg>

        '''
    }
    icone_padrao = '''
        <svg xmlns="http://www.w3.org/2000/svg" class="icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.5">
            <path stroke-linecap="round" stroke-linejoin="round"
                d="M13.5 10.5L12 12m0 0-1.5 1.5m1.5-1.5 1.5-1.5M12 12l-1.5-1.5m-1.5-1.5a3.75 3.75 0 115.303 5.303l-.978.977a3.75 3.75 0 11-5.303-5.303l.978-.977z"/>
        </svg>
    '''


    links_formatados = []
    for item in links:
        if ':' in item:
            nome, link = item.split(':', 1)
            nome = nome.strip()
            link = link.strip()
            icone = icones.get(nome, icone_padrao)
            links_formatados.append({'nome': nome, 'link': link, 'icone': icone})

    return links_formatados

@app.context_processor
def inject_request():
    return dict(request=request)

@app.route('/')
def info_page():
    file_path = 'rola.xlsx'
    df_link = pd.read_excel(file_path, header=None)
    info_essenciais = df_link.iloc[4:9, 1].dropna().tolist()
    links = df_link.iloc[10:14, 1].dropna().tolist()
    df_videos = pd.read_excel('videos_ragnarok.xlsx')
    titulos = df_videos.iloc[0:,0].dropna().tolist()
    thumb = df_videos.iloc[0:,1].dropna().tolist()
    data_publicacao = df_videos.iloc[0:,2].dropna().tolist()
    video_url = df_videos.iloc[0:,3].dropna().tolist()
    

    data_videos = list(zip(titulos,thumb,data_publicacao,video_url))
    
    videos_recentes = pd.read_excel('videos_ragnarok_unidos.xlsx',header=None)

    canal = videos_recentes.iloc[1:,1].dropna().tolist()
    titulo = videos_recentes.iloc[1:,2].dropna().tolist()
    thumb_recentes = videos_recentes.iloc[1:,3].dropna().tolist()
    data_publicacao_recentes = videos_recentes.iloc[1:,4].dropna().tolist()
    url = videos_recentes.iloc[1:,5].dropna().tolist()

    data_videos_recentes = list(zip(titulo,thumb_recentes,data_publicacao_recentes,url,canal))

    video_mais_recente = max(
    (item for item in data_videos_recentes if item[4] == "Ragnarokonlineoficial"),
    key=lambda x: x[2],
    default=None  
    )

    videos_outros_canais = [v for v in data_videos_recentes if v[4] != "Ragnarokonlineoficial"]

    mais_recentes_por_canal = {}
    for v in videos_outros_canais:
        canal = v[4]
        if canal not in mais_recentes_por_canal or v[2] > mais_recentes_por_canal[canal][2]:
            mais_recentes_por_canal[canal] = v

    videos_secundarios = sorted(mais_recentes_por_canal.values(), key=lambda x: x[2], reverse=True)

    videos_recentes_secundarios = pd.read_excel('videos_recentes.xlsx',header=None)


    canal = videos_recentes_secundarios.iloc[1:,0].dropna().tolist()
    titulo = videos_recentes_secundarios.iloc[1:,1].dropna().tolist()
    thumb_recentes = videos_recentes_secundarios.iloc[1:,2].dropna().tolist()
    data_publicacao_recentes = videos_recentes_secundarios.iloc[1:,3].dropna().tolist()
    url = videos_recentes_secundarios.iloc[1:,4].dropna().tolist()

    videos_recentes_secundarios_zip = list(zip(titulo,thumb_recentes,data_publicacao_recentes,url,canal))

    video_encontrado = next((video for video in videos_recentes_secundarios_zip if video[4].lower() == "ragnarok online latam".lower()), None)
    
    def transformar_para_embed(url):
        if "youtube.com/watch?v=" in url:
            return url.replace("watch?v=", "embed/")
        return url
    
    video_mais_recente = (
    "Ragnarok Online LATAM",
    "New exclusive costumes now available in the shop!",
    "https://i.ytimg.com/vi/kyl_YyzXXtU/maxresdefault.jpg",
    "https://www.youtube.com/embed/kyl_YyzXXtU",
    "2025-06-06T00:00:52Z"
    )

    rank_tiers = df_link.iloc[3:, 9].dropna().tolist()
    rank_classes = df_link.iloc[3:, 10].dropna().tolist()

    rank_data = list(zip(rank_tiers, rank_classes))

    sugestao_cash = df_link.iloc[31:35, 4].dropna().tolist()
    sugestao_cash_valor = df_link.iloc[31:35, 5].dropna().tolist()

    sugestao_cash_classes = df_link.iloc[38:43, 4].dropna().tolist()
    sugestao_cash_classes_valor = df_link.iloc[38:43, 5].dropna().tolist()


    cash = list(zip(sugestao_cash, sugestao_cash_valor))
    cash_para_classes = list(zip(sugestao_cash_classes, sugestao_cash_classes_valor))




    return render_template(
        'index.html',
        info=info_essenciais,
        links=carregar_links(),
        rank_data = rank_data,
        data_videos = data_videos,
        streamers = stream_cache,
        video_mais_recente = video_mais_recente,
        videos_secundarios = videos_recentes_secundarios_zip,
        sugestao_cash = cash,
        sugestao_cash_classes = cash_para_classes
        )

@app.route('/classes')
def classes_page():
    df = pd.read_excel('rola.xlsx', header=None)
    classes = df.iloc[1:27, 4].dropna().tolist()
    builds = df.iloc[1:29, 5].dropna().tolist()
    print(classes)
    print(builds)

    wb = load_workbook('rola.xlsx')
    ws = wb["INFORMAÇÕES"]
        
    links_classes = []
    for row in ws.iter_rows(min_row=33, min_col=2, max_col=2):
            cell = row[0]
            if cell.hyperlink:
                links_classes.append(cell.hyperlink.target)
            elif cell.value:
                links_classes.append(cell.value)

    df_guias = pd.read_excel('guias_classes.xlsx',header=None)
    df_guias.columns = ['id', 'channel_name', 'title', 'thumbnail', 'video_url', 'classe']
    
    classe_map = {
    "Cavaleiros Rúnicos": "Cavaleiro Rúnico",
    "Guardião Real": "Guardião Real",
    "Feiticeiro": "Feiticeiro",
    "Sentinela": "Sentinela",
    "Sicário": "Sicário",
    "Arcanos": "Arcano",
    "Arcebispos": "Arcebispo",
    "Renegado": "Renegado",
    "Shura": "Sura",
    "Mecânico": "Mecânico",
    "Bioquímicos": "Bioquímico",
    "Trovadores": "Trovador",
    "Musa": "Trovador"
    }   
    def normalizar(texto):
        return str(texto).strip().lower()
    df_guias["classe_normalizada"] = df_guias["classe"].apply(normalizar)
    

    class_builds = [
        {"classe": classes[1], "builds": [builds[1], builds[2]], "link": "","imagem": "assets/classes/feiticeiro.png","imagem_sentado": "assets/classes/sentados/feiticeiro.gif","imagem_andando": "assets/classes/andando/feiticeiro.gif"}, #Feiticeiro
        {"classe": classes[2], "builds": [builds[3]], "link": links_classes[4],"imagem": "assets/classes/sentinela.png","imagem_sentado": "assets/classes/sentados/sentinela.gif","imagem_andando": "assets/classes/andando/sentinela.gif"},# Sentinela
        {"classe": classes[3], "builds": [builds[4]], "link": links_classes[5],"imagem": "assets/classes/sicario.png","imagem_sentado": "assets/classes/sentados/sicario.gif","imagem_andando": "assets/classes/andando/sicario.gif"},# Sicario
        {"classe": classes[4], "builds": [builds[5]], "link": links_classes[6],"imagem": "assets/classes/arcano.png","imagem_sentado": "assets/classes/sentados/arcano.gif","imagem_andando": "assets/classes/andando/arcano.gif"},# Arcano
        {"classe": classes[5], "builds": [builds[6], builds[7]], "link": "","imagem": "assets/classes/arcebispo.png","imagem_sentado": "assets/classes/sentados/arcebispo.gif","imagem_andando": "assets/classes/andando/arcebispo.gif"},# Arcebispo
        {"classe": classes[6], "builds": [builds[9], builds[8]], "link": links_classes[3],"imagem": "assets/classes/renegado.png","imagem_sentado": "assets/classes/sentados/renegado.gif","imagem_andando": "assets/classes/andando/renegado.gif"},# Renegado
        {"classe": classes[7], "builds": [builds[11], builds[10]], "link": "","imagem": "assets/classes/sura.png","imagem_sentado": "assets/classes/sentados/shura.gif","imagem_andando": "assets/classes/andando/shura.gif"},# Shura
        {"classe": classes[8], "builds": [builds[12]], "link": links_classes[1],"imagem": "assets/classes/cavaleiro_runico.png","imagem_sentado": "assets/classes/sentados/cavaleiro_runico.gif","imagem_andando": "assets/classes/andando/cavaleiro_runico.gif"}, # Cavaleiros Rúnicos
        {"classe": classes[9], "builds": [builds[14], builds[13]], "link": links_classes[2],"imagem": "assets/classes/guardioes_reais.png","imagem_sentado": "assets/classes/sentados/guardiao_real.gif","imagem_andando": "assets/classes/andando/guardiao_real.gif"},# Guardião Real
        {"classe": classes[10], "builds": [builds[15]], "link": "","imagem": "assets/classes/mecanico.png","imagem_sentado": "assets/classes/sentados/mecanico.gif","imagem_andando": "assets/classes/andando/mecanico.gif"},# Mecânico
        {"classe": classes[11], "builds": [builds[16]], "link": "","imagem": "assets/classes/genetico.png","imagem_sentado": "assets/classes/sentados/bioquimico.gif","imagem_andando": "assets/classes/andando/bioquimico.gif"},# Bioquímicos
        {"classe": classes[12], "builds": [builds[17], builds[18]], "link": links_classes[7],"imagem": "assets/classes/trovador.png","imagem_sentado": "assets/classes/sentados/trovador.gif","imagem_andando": "assets/classes/andando/trovador.gif"},# Trovadores
        {"classe": classes[13], "builds": [builds[19], builds[20]], "link": links_classes[7],"imagem": "assets/classes/musa.png","imagem_sentado": "assets/classes/sentados/musa.gif","imagem_andando": "assets/classes/andando/musa.gif"}, # Musa
    ]
    

    def transformar_para_embed(url):
        if "youtube.com/watch?v=" in url:
            return url.replace("watch?v=", "embed/")
        return url


    for item in class_builds:
        nome_original = item["classe"]
        nome_para_buscar = classe_map.get(nome_original, nome_original)  
        classe_normalizada = normalizar(nome_para_buscar)

        guias_filtrados = df_guias[df_guias["classe_normalizada"] == classe_normalizada]

        lista_guias = []
        for _, row in guias_filtrados.iterrows():
            guia = {
                "id": row["id"],
                "channel_name": row["channel_name"],
                "title": row["title"],
                "thumbnail": row["thumbnail"],
                "video_url": row["video_url"]
            }
            lista_guias.append(guia)
        item["guias"] = lista_guias

    
    colunas = ['channel_name', 'title', 'thumbnail', 'video_url','classe']
    df_guias_gerais = pd.read_excel('guias_gerais.xlsx', header=None, names=colunas)


    df_guias_gerais['embed_url'] = df_guias_gerais['video_url'].apply(transformar_para_embed)
    guias_gerais = df_guias_gerais[df_guias_gerais['classe'] == 'Geral']
    guias_gerais = df_guias_gerais[df_guias_gerais['classe'] == 'Geral'].to_dict(orient='records')


    

    links = carregar_links()
    return render_template('classes.html', 
                           class_builds=class_builds,
                           links = links,
                           streamers = stream_cache,
                           guias_gerais  = guias_gerais
                           )

@app.route('/rank')
def rank_page():
    df = pd.read_excel('rola.xlsx', header=None)
    rank_tiers = df.iloc[3:, 8].dropna().tolist()
    rank_classes = df.iloc[3:, 9].dropna().tolist()

    links = carregar_links()

    rank_data = list(zip(rank_tiers, rank_classes))
    return render_template('rank.html', 
                           rank_data=rank_data,
                           links= links,
                           streamers = stream_cache
                           )

@app.route('/rotas')
def rotas_page():
    df_melee = pd.read_excel('rola.xlsx', sheet_name='ROTAS MELEE + DICAS', header=None)
    melee_raw = df_melee.iloc[2:,1].dropna().tolist()
    quest_melee =  df_melee.iloc[1:,3].dropna().tolist() 
    builds_melee = df_melee.iloc[1:,4].dropna().tolist() 

    print(builds_melee)

    df_ranged = pd.read_excel('rola.xlsx', sheet_name='ROTAS RANGED + DICAS', header=None)
    ranged_raw = df_ranged.iloc[2:,1].dropna().tolist()
    quest_ranged =  df_ranged.iloc[1:,3].dropna().tolist() 
    builds_ranged = df_ranged.iloc[1:,4].dropna().tolist() 

    #ROTAS MELEE
    rotas_melee = []
    for linha in melee_raw:
        linha = str(linha).strip()

        match = re.match(r"^(\d+~\d+|\d+\+?)\s*(.*)", linha)

        if match:
            nivel = match.group(1).strip()
            local = match.group(2).strip()
        else:
            nivel = ""
            local = linha

        rotas_melee.append((nivel, local))
    
    rotas_melee_filtrado = []

    for nivel, local in rotas_melee:
        if "Dano AoE:" in local:
            continue 

        if "Quests Éden para UP" in local:
            break 

        rotas_melee_filtrado.append((nivel, local))

    array_quest_melee = [
        (quest_melee[1]),
        (quest_melee[2]),
        (quest_melee[3]),
        (quest_melee[4]),
        (quest_melee[5]),
        (quest_melee[6]),
        (quest_melee[8]),
        (quest_melee[9]),
        (quest_melee[10])
    ]

    array_builds_melee = [
         (quest_melee[14], builds_melee[4]),
         (quest_melee[15], builds_melee[5]),
         (quest_melee[16], builds_melee[6]),
         (quest_melee[17], builds_melee[7]),
         (quest_melee[18], builds_melee[8]),
         (quest_melee[19], builds_melee[9]),
         (quest_melee[20], builds_melee[10]),
    ]

    array_builds_formatado =[]
   
    for quest, build in array_builds_melee:
        quest = str(quest).strip()
        build = str(build).strip()

        quest = re.sub(r"^[('\"\s]+|[)'\"]+$", "", quest)
        build = re.sub(r"^[('\"\s]+|[)'\"]+$", "", build)

        array_builds_formatado.append((quest, build))
    
    rotas_ranged = []
    for linha in ranged_raw:
        linha = str(linha).strip()

        match = re.match(r"^(\d+~\d+|\d+\+?)\s*(.*)", linha)

        if match:
            nivel = match.group(1).strip()
            local = match.group(2).strip()
        else:
            nivel = ""
            local = linha

        rotas_ranged.append((nivel, local))

    rotas_ranged_filtrado = []

    for nivel, local in rotas_ranged:
        if "Dano AoE:" in local:
            continue  

        if "Quests Éden para UP" in local:
            break  
        rotas_ranged_filtrado.append((nivel, local))   

    array_quest_ranged = [
        ("Equipamentos " + quest_ranged[1]),
        (quest_ranged[2]),
        (quest_ranged[3]),
        (quest_ranged[4]),
        (quest_ranged[5]),
        (quest_ranged[6]),
        (quest_ranged[8]),
        (quest_ranged[9]),
        (quest_ranged[10])
    ]

    array_builds_ranged = [
         (quest_ranged[14], builds_ranged[4]),
         (quest_ranged[15], builds_ranged[5]),
         (quest_ranged[16], builds_ranged[6]),
         (quest_ranged[17], builds_ranged[7]),
         (quest_ranged[18], builds_ranged[8]),
    ]
    array_builds_ranged_formatado =[]
   
    for quest, build in array_builds_ranged:
    # Converte para string e remove espaços
        quest = str(quest).strip()
        build = str(build).strip()

    # Remove parênteses e aspas se houver
        quest = re.sub(r"^[('\"\s]+|[)'\"]+$", "", quest)
        build = re.sub(r"^[('\"\s]+|[)'\"]+$", "", build)

    # Adiciona como tupla limpa
        array_builds_ranged_formatado.append((quest, build))
   
    links = carregar_links()

    return render_template(
        'rotas.html',
        rotas_melee=rotas_melee_filtrado,
        array_quest_melee=array_quest_melee,
        array_builds_formatado = array_builds_formatado,
        rotas_ranged=rotas_ranged_filtrado,
        array_quest_ranged=array_quest_ranged,
        array_builds_ranged_formatado = array_builds_ranged_formatado,
        links=links,
        streamers = stream_cache
    )

@app.route('/items')
def items_page():
    file_path = 'rola.xlsx'
    df = pd.read_excel(file_path, sheet_name='TABELA ITENS', header=None)
    df_items = pd.read_excel(file_path, sheet_name='ITENS NÃO VENDER', header=None)

    items = df.iloc[3:,1]
    utilizar = df.iloc[3:,2]


    
    arquivo = ["item_db_usable.yml","item_db_equip.yml","item_db_etc.yml"]
    items_n_vender_raw_3 = df_items.iloc[3:,6].dropna().tolist()
    items_n_vender_raw_3_2 = df_items.iloc[3:,7].dropna().tolist()
    items_n_vender_raw_3_2_3 = df_items.iloc[3:,8].dropna().tolist()

    items_n_vender_raw_4 = df_items.iloc[3:,12].dropna().tolist()
    items_n_vender_raw_4_2 = df_items.iloc[3:,14].dropna().tolist()



    items_n_vender_raw_5 = df_items.iloc[3:,20].dropna().tolist()
    items_n_vender_raw_6 = df_items.iloc[3:,24].dropna().tolist()
    items_n_vender_raw_6_2 = df_items.iloc[3:,26].dropna().tolist()
    items_n_vender_raw_6_3 = df_items.iloc[3:,28].dropna().tolist()
    items_n_vender_raw_6_4 = df_items.iloc[3:,30].dropna().tolist()
    items_n_vender_raw_6_5 = df_items.iloc[3:,32].dropna().tolist()



    items_n_vender_raw_7 = df_items.iloc[3:,36].dropna().tolist()
    items_n_vender_raw_7_2 = df_items.iloc[3:,37].dropna().tolist()
    items_n_vender_raw_7_3 = df_items.iloc[3:,38].dropna().tolist()
    items_n_vender_raw_7_4 = df_items.iloc[3:,39].dropna().tolist()



    items_n_vender_raw_8 = df_items.iloc[3:,43].dropna().tolist()
    items_n_vender_raw_8_2 = df_items.iloc[3:,44].dropna().tolist()
    items_n_vender_raw_8_3 = df_items.iloc[3:,45].dropna().tolist()
    items_n_vender_raw_8_4 = df_items.iloc[3:,46].dropna().tolist()



    items_n_vender_raw_9 = df_items.iloc[12:,45].dropna().tolist()





    def normalizar(texto):
        if texto is None:
            return ""
        return unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('ASCII').strip().lower()
    
    def buscar_id_por_nome_em_arquivos(nome_procurado, arquivo):
        nome_procurado_norm = normalizar(nome_procurado)

        for caminho_arquivo in arquivo:
            try:
                with open(caminho_arquivo, 'r', encoding='latin-1') as file:
                    bloco = []
                    for linha in file:
                        if linha.strip().startswith('- Id:'):
                            if bloco:
                                for l in bloco:
                                    if l.strip().startswith("Name:"):
                                        nome_linha = l.split(":", 1)[1].strip()
                                        if normalizar(nome_linha) == nome_procurado_norm:
                                            for b in bloco:
                                                if 'Id:' in b:
                                                    return b.split(':', 1)[1].strip()
                            bloco = [linha]
                        else:
                            bloco.append(linha)

                    # Verifica o último bloco
                    for l in bloco:
                        if l.strip().startswith("Name:"):
                            nome_linha = l.split(":", 1)[1].strip()
                            if normalizar(nome_linha) == nome_procurado_norm:
                                for b in bloco:
                                    if 'Id:' in b:
                                        return b.split(':', 1)[1].strip()

            except Exception as e:
                print(f"Erro ao processar {caminho_arquivo}: {e}")

        return None
    imagens = []

    # Loop para percorrer os itens e montar a URL da imagem com base no ID
    for item in items:
        item_id = buscar_id_por_nome_em_arquivos(item, arquivo)
        if item_id:
            imagem_url = f"https://www3.worldrag.com/database/media/item/{item_id}.gif"
            imagens.append(imagem_url)
        else:
            imagens.append("imagem_nao_encontrada")

    

    tabela_items = list(zip(items,utilizar,imagens))

    items_importantes = df.iloc[9:20,4].dropna().tolist()
    items_importantes_local = df.iloc[9:20,5].dropna().tolist()

   


    imagens_items_importantes = []
    
    for item in items_importantes:
        id_str = buscar_id_por_nome_em_arquivos(item, arquivo)
        if id_str:
            imagem_url = f"https://www3.worldrag.com/database/media/item/{id_str}.gif"
            imagens_items_importantes.append(imagem_url)
        else:
            imagens_items_importantes.append("imagem_nao_encontrada")

    tabela_items_importantes = list(zip(items_importantes,items_importantes_local,imagens_items_importantes))



    ################  items para o futuro ###############################

    item_guardar_intancia = df.iloc[23:,4].dropna().tolist()
    item_guardar = df.iloc[23:,5].dropna().tolist()
    item_guardar_moedas = df.iloc[23:,6].dropna().tolist()
    

    array_items_guardar = [
        (item_guardar_intancia[0],item_guardar[0],item_guardar_moedas[0]), # Altar do Selo
        (item_guardar_intancia[1],item_guardar[1],item_guardar_moedas[1],item_guardar[2],
        item_guardar_moedas[2],item_guardar[3],item_guardar_moedas[3],item_guardar[4],item_guardar_moedas[4],
        item_guardar[5],item_guardar_moedas[5],item_guardar[6],item_guardar_moedas[6],), # Invasão ao Aeroplano
        (item_guardar_intancia[2],item_guardar[7],item_guardar_moedas[7],item_guardar[8],item_guardar_moedas[8],
        item_guardar[9],item_guardar_moedas[9],item_guardar[10],item_guardar_moedas[10],item_guardar[11],item_guardar_moedas[11],item_guardar[12],
        item_guardar_moedas[12],item_guardar[13],item_guardar_moedas[13],item_guardar[14],item_guardar_moedas[14]), #Maldição de Glast Heim
        (item_guardar_intancia[3],item_guardar[15],item_guardar_moedas[15],item_guardar[16],item_guardar_moedas[16],item_guardar[17],item_guardar_moedas[17],
        item_guardar[18],item_guardar_moedas[18],item_guardar[19],item_guardar_moedas[19],item_guardar[20],item_guardar_moedas[20]),# Torre do Demônio
        (item_guardar_intancia[4],item_guardar[21],item_guardar_moedas[21],item_guardar[22],item_guardar_moedas[22]), # Sala Final
        (item_guardar_intancia[5],item_guardar[23],item_guardar_moedas[23]), # Ilha Bios
        (item_guardar_intancia[6],item_guardar[24],item_guardar_moedas[24]), # Caverna de Mors
    ]
    array_items_guardar_tabela = []


    array_link_instancias = [
        "https://browiki.org/wiki/Altar_do_Selo",
        "https://browiki.org/wiki/Invas%C3%A3o_ao_Aeroplano",
        "https://browiki.org/wiki/Maldi%C3%A7%C3%A3o_de_Glast_Heim",
        "https://browiki.org/wiki/Torre_do_Dem%C3%B4nio",
        "https://browiki.org/wiki/Sala_Final",
        "https://browiki.org/wiki/Ilha_Bios",
        "https://browiki.org/wiki/Caverna_de_Mors"
    ]
    # Iterar por cada bloco
    for bloco in array_items_guardar:
        instancia = bloco[0]
        dados = bloco[1:]
        
        # pares de item/moeda
        for i in range(0, len(dados), 2):
            item = dados[i]
            moeda = dados[i + 1]
            array_items_guardar_tabela.append((instancia, item, moeda))
   

    tabela_formatada = []
    instancia_anterior = None

    for instancia, item, moeda in array_items_guardar_tabela:
            if instancia != instancia_anterior:
                tabela_formatada.append((instancia, item, moeda))
                instancia_anterior = instancia
            else:
                tabela_formatada.append(("", item, moeda))
 

    dados_com_links = []
    link_index = 0
    for nome, descricao, quantidade in tabela_formatada:
        if nome.strip():  # se nome não está vazio
            link = array_link_instancias[link_index] if link_index < len(array_link_instancias) else None
            link_index += 1
        else:
            link = " "

        dados_com_links.append((nome, descricao, quantidade, link))

    for item in dados_com_links:
        print(item)

        ###################  items n vender ###############
    
    array_items_n_vender = [
    {"classe": items_n_vender_raw_3[0].replace(" Nada", "").strip(), "items": [], "imagem": "assets/classes/espadachin.png"}, #Espadachin
    {"classe": items_n_vender_raw_3[1], "items": [
        [
            items_n_vender_raw_3[2],
            items_n_vender_raw_3[3],
            items_n_vender_raw_3[4],
            items_n_vender_raw_3[5],
            items_n_vender_raw_3[6],
            items_n_vender_raw_3[7]
        ],
        [
            items_n_vender_raw_3_2_3[0],
            items_n_vender_raw_3_2_3[1],
            items_n_vender_raw_3_2_3[2],
            items_n_vender_raw_3_2_3[3],
            items_n_vender_raw_3_2_3[4],
            items_n_vender_raw_3_2_3[5]
        ]
    ],
    "imagem": "assets/classes/cavaleiro.png"
    }, #Cavaleiro
    {"classe": items_n_vender_raw_3[8], "items": [
        [
            items_n_vender_raw_3[9],
            items_n_vender_raw_3[10]
        ],
        [
            items_n_vender_raw_3_2_3[6],
            items_n_vender_raw_3_2_3[7]
        ]
    ],
    "imagem": "assets/classes/templario.png"
    }, #Templario
    {"classe": items_n_vender_raw_3[11].replace(" Nada", "").strip(), "items": [], "imagem": "assets/classes/cavaleiro_runico.png"}, #Cavaleiro Runico
    {"classe": items_n_vender_raw_3[12], "items": [
        [items_n_vender_raw_3[13]]
    ],
    "imagem": "assets/classes/guardioes_reais.png"
    },#Guardiao Real
    {"classe": items_n_vender_raw_4[0].replace(" Nada", "").strip(), "items": [],"imagem": "assets/classes/mago.png"}, #Mago
    {"classe": items_n_vender_raw_4[1].replace(" Nada", "").strip(), "items": [
        [
            items_n_vender_raw_4[2],
            items_n_vender_raw_4[3],
            items_n_vender_raw_4[4]
        ],
        [
            items_n_vender_raw_4_2[0],
            items_n_vender_raw_4_2[1],
            items_n_vender_raw_4_2[2],
            items_n_vender_raw_4_2[3]
        ]
    ],
    "imagem": "assets/classes/bruxo.png"
    }, #Bruxo
    {"classe": items_n_vender_raw_4[5].replace(" Nada", "").strip(), "items": [
        [
            items_n_vender_raw_4[6],
            items_n_vender_raw_4[7]
        ],
        [
            items_n_vender_raw_4_2[4],
            items_n_vender_raw_4_2[5],
            items_n_vender_raw_4_2[6]
        ]
    ],
    "imagem": "assets/classes/sabio.png"
    }, #Sabio
    {"classe": items_n_vender_raw_4[20].replace(" Nada", "").strip(), "items": [],"imagem": "assets/classes/arcano.png"}, #Arcano
    {"classe": items_n_vender_raw_4[21].replace(" Nada", "").strip(), "items": [
        [
            items_n_vender_raw_4[22],
            items_n_vender_raw_4[23],
            items_n_vender_raw_4[24],
            items_n_vender_raw_4[25]
        ]
    ],
    "imagem": "assets/classes/feiticeiro.png"
    },#Feiticeiro
    {"classe": items_n_vender_raw_5[0].replace(" Nada", "").strip(), "items": [],"imagem": "assets/classes/gatuno.png"}, #Gatuno
    {"classe": items_n_vender_raw_5[1].replace(" Nada", "").strip(), "items": [],"imagem": "assets/classes/mercenario.png"},#Mercenario
    {"classe": items_n_vender_raw_5[2].replace(" Nada", "").strip(), "items": [
        [
            items_n_vender_raw_5[3],
            items_n_vender_raw_5[4],
            items_n_vender_raw_5[5],
            items_n_vender_raw_5[6]
        ]
    ],
    "imagem": "assets/classes/arruaceiro.png"
    },#Arruaceiro
    {"classe": items_n_vender_raw_5[7].replace(" Nada", "").strip(), "items": [],"imagem": "assets/classes/sicario.png"},#Sicario
    {"classe": items_n_vender_raw_5[8].replace(" Nada", "").strip(), "items": [],"imagem": "assets/classes/renegado.png"},#Renegado
    {"classe": items_n_vender_raw_6[0].replace(" Nada", "").strip(), "items": [],"imagem": "assets/classes/mercador.png"},#Mercador
    {"classe": items_n_vender_raw_6[1].replace(" Nada", "").strip(), "items": [
        [
            items_n_vender_raw_6[2],
            items_n_vender_raw_6[3],
            items_n_vender_raw_6[4],
            items_n_vender_raw_6[5]
        ],
        [
            items_n_vender_raw_6_2[0],
            items_n_vender_raw_6_2[1],
            items_n_vender_raw_6_2[2],
            items_n_vender_raw_6_2[3]
        ],
        [
            items_n_vender_raw_6_3[0],
            items_n_vender_raw_6_3[1],
            items_n_vender_raw_6_3[2],
            items_n_vender_raw_6_3[3]
        ],
        [
            items_n_vender_raw_6_4[0],
            items_n_vender_raw_6_4[1],
            items_n_vender_raw_6_4[2],
            items_n_vender_raw_6_4[3]
        ],
        [
            items_n_vender_raw_6_5[0],
            items_n_vender_raw_6_5[1],
            items_n_vender_raw_6_5[2],
            items_n_vender_raw_6_5[3]
        ]
    ],
    "imagem": "assets/classes/ferreiro.png"},#Ferreiro
    {"classe": items_n_vender_raw_6[6].replace(" Nada", "").strip(), "items": [
        [
            items_n_vender_raw_6[7],
            items_n_vender_raw_6[8],
            items_n_vender_raw_6[9],
            items_n_vender_raw_6[10],
            items_n_vender_raw_6[11],
            items_n_vender_raw_6[12],
            items_n_vender_raw_6[13],
            items_n_vender_raw_6[14],
            items_n_vender_raw_6[15],
            items_n_vender_raw_6[16]
        ],
        [
            items_n_vender_raw_6_2[4],
            items_n_vender_raw_6_2[5]
        ],
        [
            items_n_vender_raw_6_2[6]
        ]
    ],
    "imagem": "assets/classes/alquimista.png"
    },#Alquimista
    {"classe": items_n_vender_raw_6[17].replace(" Nada", "").strip(), "items": [],"imagem": "assets/classes/mecanico.png"},#Mecanico
    {"classe": items_n_vender_raw_6[18].replace(" Nada", "").strip(), "items": [],"imagem": "assets/classes/genetico.png"},#Genetico
    {"classe": items_n_vender_raw_7[0].replace(" Nada", "").strip(), "items": [],"imagem": "assets/classes/novico.png"},#Novico
    {"classe": items_n_vender_raw_7[1].replace(" Nada", "").strip(), "items": [],"imagem": "assets/classes/sacerdote.png"},#Sacerdote
    {"classe": items_n_vender_raw_7[2].replace(" Nada", "").strip(), "items": [
        [
            "Possibilidade 1",
            items_n_vender_raw_7_2[1],
            items_n_vender_raw_7_2[2],
            items_n_vender_raw_7_2[3],
            items_n_vender_raw_7_2[4],
            items_n_vender_raw_7_2[5],
            items_n_vender_raw_7_2[6],
            items_n_vender_raw_7_2[7],

        ], 
        [
            "Possibilidade 2",
            items_n_vender_raw_7_3[1],
            items_n_vender_raw_7_3[2],
            items_n_vender_raw_7_3[3],
            items_n_vender_raw_7_3[4],
            items_n_vender_raw_7_3[5],
            items_n_vender_raw_7_3[6],
        ],
        [
            "Possibilidade 3",
            items_n_vender_raw_7_4[1],
            items_n_vender_raw_7_4[2],
            items_n_vender_raw_7_4[3],
            items_n_vender_raw_7_4[4],
            items_n_vender_raw_7_4[5],
            items_n_vender_raw_7_4[6],
        ]
    ],
    "imagem": "assets/classes/monge.png"
    },#Monge
    {"classe": items_n_vender_raw_7[11].replace(" Nada", "").strip(), "items": [],"imagem": "assets/classes/arcebispo.png"},#Arcebispo
    {"classe": items_n_vender_raw_7[12].replace(" Nada", "").strip(), "items": [],"imagem": "assets/classes/sura.png"},#Sura
    {"classe": items_n_vender_raw_8[0].replace(" Nada", "").strip(), "items": [],"imagem": "assets/classes/arqueiro.png"},#arqueiro
    {"classe": items_n_vender_raw_8[1].replace(" Nada", "").strip(), "items": [
        [
            items_n_vender_raw_8[2],
            items_n_vender_raw_8[3],
        ],[    
            items_n_vender_raw_8_2[0],
            items_n_vender_raw_8_2[1],
        ],
        [
            items_n_vender_raw_8_3[0],
            items_n_vender_raw_8_3[1],
        ],[
              items_n_vender_raw_8_4[0],
            items_n_vender_raw_8_4[1],
        ],
        [
            items_n_vender_raw_8[4],
            items_n_vender_raw_8[5],
        ],
        [
            items_n_vender_raw_8_2[2],
            items_n_vender_raw_8_2[3],
        ],
        [
            items_n_vender_raw_8_3[2],
            items_n_vender_raw_8_3[3],
        ],
        # [
        #     items_n_vender_raw_8_2[2],
        #     items_n_vender_raw_8_2[3],
        #     items_n_vender_raw_8_2[4],
        # ]

    ],"imagem": "assets/classes/caçador.png"},#Caçador
    {"classe": items_n_vender_raw_8[6].replace(" Nada", "").strip(), "items": [
        [
            items_n_vender_raw_8[7],
            items_n_vender_raw_8[8],
            items_n_vender_raw_8[9],
            items_n_vender_raw_8[10],
            items_n_vender_raw_8[11],
            items_n_vender_raw_8[12],
            items_n_vender_raw_8[13],
            items_n_vender_raw_8[14],
            items_n_vender_raw_8[15]
        ]],
     "imagem": "assets/classes/bardo.png"},#Bardo
    {"classe": items_n_vender_raw_8[16].replace(" Nada", "").strip(), "items": [],"imagem": "assets/classes/sentinela.png"},#Sentinela
    {"classe": items_n_vender_raw_8[17].replace(" Nada", "").strip(), "items": [],"imagem": "assets/classes/trovador.png"},#Trovador
    {"classe": items_n_vender_raw_9[0], "items": [
        [
            items_n_vender_raw_9[1],
            items_n_vender_raw_9[2],
            items_n_vender_raw_9[3],
            items_n_vender_raw_9[4],
            items_n_vender_raw_9[5],
        ],
        [  
            items_n_vender_raw_9[6],
            items_n_vender_raw_9[7],
            items_n_vender_raw_9[8],
        ],
        [
            items_n_vender_raw_9[9],
            items_n_vender_raw_9[10],
            items_n_vender_raw_9[11],
            items_n_vender_raw_9[12],
            items_n_vender_raw_9[13],
            items_n_vender_raw_9[14]
        ]
    ],
    "imagem": "assets/classes/odalisca.png"
    },#Odalistica
    {"classe": items_n_vender_raw_9[15].replace(" Nada", "").strip(), "items": [],"imagem": "assets/classes/musa.png"}#Musa
]
    
    links = carregar_links()
    return render_template('items.html', 
                           tabela_items=tabela_items,
                           tabela_items_importantes=tabela_items_importantes,
                           array_items_guardar=dados_com_links,
                           array_items_n_vender=array_items_n_vender,
                           links=links,
                           streamers = stream_cache
                           )

@app.route('/monstros')
def monstros_page():
    file_path = 'rola.xlsx'
    df = pd.read_excel(file_path, sheet_name='Monstros + Drop + XP', header=None)
    df_monstro_novo = pd.read_excel('monstros_e_drops.xlsx',header=None)



    caminho_arquivos = 'itens.xlsx'
 

    def buscar_ids_itens_para_monstro(itens_str, dict_ids):
        nomes_itens = [item.strip() for item in itens_str.split(',') if item.strip()]
        resultado = []

        for nome in nomes_itens:
            id_encontrado = dict_ids.get(nome)
            resultado.append({
                "nome": nome,
                "id": id_encontrado
            })

        return resultado

    def carregar_ids_em_memoria(caminho_arquivo):
        dicionario = {}

        try:
            df = pd.read_excel(caminho_arquivo, header=None)

            for _, row in df.iterrows():
                id_item = str(row[0]).strip()
                nome = str(row[1]).strip()
                if nome and id_item:
                    dicionario[nome] = id_item

        except Exception as e:
            print(f"Erro ao processar {caminho_arquivo}: {e}")

        return dicionario


    df_monstros_novo = pd.DataFrame({
        "id": df_monstro_novo.iloc[1:, 0],
        "nome": df_monstro_novo.iloc[1:, 1],
        "level": df_monstro_novo.iloc[1:, 2],
        "hp": df_monstro_novo.iloc[1:, 3],
        "exp_base": df_monstro_novo.iloc[1:, 4],
        "exp_job": df_monstro_novo.iloc[1:, 5],
        "itens": df_monstro_novo.iloc[1:, 7],
        "mapa": df_monstro_novo.iloc[1:, 9]
    })

    df_agrupado = df_monstros_novo.groupby("nome", as_index=False).agg({
        "id": "first",
        "level": "first",
        "hp": "first",
        "exp_base": "first",
        "exp_job": "first",
        "itens": "first",
        "mapa": lambda x: ', '.join(sorted(set(str(i).strip() for i in x if pd.notna(i))))
    })

    dicionario_ids = carregar_ids_em_memoria(caminho_arquivos)

    df_agrupado["itens_com_id"] = df_agrupado["itens"].apply(
        lambda x: buscar_ids_itens_para_monstro(x, dicionario_ids)
    )
    
    data = df_agrupado.to_dict(orient='records')

    links = carregar_links()

    return render_template('monstros.html', data=data,links=links,streamers=stream_cache)

@app.route("/timer")
def timer_page():

    file_path = 'rola.xlsx'

    df_xp_segunda_classe = pd.read_excel(file_path, sheet_name='Tabela de XP Segunda Classe', header=None)
    df_xp_terceira_classe = pd.read_excel(file_path, sheet_name='Tabela de XP Terceira Classe', header=None)
    data =""

    bloco1 = df_xp_segunda_classe.iloc[2:, 1:4]
    bloco2 = df_xp_segunda_classe.iloc[2:, 5:8]
    bloco3 = df_xp_segunda_classe.iloc[2:, 9:12]

    for bloco in [bloco1, bloco2, bloco3]:
        bloco.columns = ['Level', 'Total EXP', 'EXP Proximo Level']

    tabela_unificada = pd.concat([bloco1, bloco2, bloco3], ignore_index=True)

    tabela_unificada = tabela_unificada.dropna(how='all')

    tabela_unificada['Level'] = pd.to_numeric(tabela_unificada['Level'], errors='coerce')
    tabela_unificada = tabela_unificada.dropna(subset=['Level'])
    tabela_unificada['Level'] = tabela_unificada['Level'].astype(int)


    df_xp_tranclasse = pd.read_excel(file_path, sheet_name='Tabela de XP Transclasse', header=None)
    
    bloco1_tranclasse = df_xp_tranclasse.iloc[2:, 1:4]
    bloco2_tranclasse = df_xp_tranclasse.iloc[2:, 5:8]
    bloco3_tranclasse = df_xp_tranclasse.iloc[2:, 9:12]


    for bloco in [bloco1_tranclasse, bloco2_tranclasse, bloco3_tranclasse]:
        bloco.columns = ['Level', 'Total EXP', 'EXP Proximo Level']

    tabela_unificada_tranclasse = pd.concat([bloco1_tranclasse, bloco2_tranclasse, bloco3_tranclasse], ignore_index=True)

    tabela_unificada_tranclasse = tabela_unificada_tranclasse.dropna(how='all')

    tabela_unificada_tranclasse['Level'] = pd.to_numeric( tabela_unificada_tranclasse['Level'], errors='coerce')
    tabela_unificada_tranclasse =  tabela_unificada_tranclasse.dropna(subset=['Level'])
    tabela_unificada_tranclasse['Level'] =  tabela_unificada_tranclasse['Level'].astype(int)




    bloco1_terceira = df_xp_terceira_classe.iloc[2:, 1:4]
    bloco2_terceira = df_xp_terceira_classe.iloc[2:, 5:8]
    bloco3_terceira = df_xp_terceira_classe.iloc[2:, 9:12]

    for bloco in [bloco1_terceira, bloco2_terceira, bloco3_terceira]:
        bloco.columns = ['Level', 'Total EXP', 'EXP Proximo Level']


    tabela_unificada_terceira = pd.concat([bloco1_terceira, bloco2_terceira, bloco3_terceira], ignore_index=True)

    tabela_unificada_terceira = tabela_unificada_terceira.dropna(how='all')

    tabela_unificada_terceira['Level'] = pd.to_numeric( tabela_unificada_terceira['Level'], errors='coerce')
    tabela_unificada_terceira =  tabela_unificada_terceira.dropna(subset=['Level'])
    tabela_unificada_terceira['Level'] =  tabela_unificada_terceira['Level'].astype(int)

    links= carregar_links()

    return render_template('timer.html', data=data,
                            tabela_unificada_segunda_classe = tabela_unificada.to_dict(orient='records'),
                            tabela_unificada_transclasse = tabela_unificada_tranclasse.to_dict(orient='records'),
                            tabela_unificada_terceira = tabela_unificada_terceira.to_dict(orient='records'),
                            links=links,
                            streamers =stream_cache
                            )
@app.route('/streamers')
def streamers_page():
    file_path = 'rola.xlsx'

    df = pd.read_excel(file_path, sheet_name='INFORMAÇÕES', header=None)


    links = carregar_links()
    return render_template('streamers.html', data=stream_cache,
                            links=links,
                            streamers = stream_cache
                            )



@app.route('/utilitarios')
def utilitarios_page():

   return render_template('utilitarios.html',links=carregar_links(),
)

@app.route('/contato&apoio')
def contato_page():

   return render_template('contato&apoio.html',links=carregar_links(),streamers = stream_cache
)



@app.route('/spots')
def melhores_spots_page():
    file_path = 'melhor_spot_por_level.xlsx'
    df = pd.read_excel(file_path, header=None)

    monstros = df.iloc[1:,1].dropna().tolist()
    base_exp = df.iloc[1:,3].dropna().tolist()
    job_exp = df.iloc[1:,4].dropna().tolist()
    quantidade = df.iloc[1:,5].dropna().tolist()
    mapa = df.iloc[1:,6].dropna().tolist()
    xp_ajustada = df.iloc[1:,7].dropna().tolist()
    nivel_jogador = df.iloc[1:,8].dropna().tolist()
    nivel_monstro = df.iloc[1:,2].dropna().tolist()

    dicionario_ids= 'monstros_e_drops.xlsx'

    def normalizar(texto):
        if not isinstance(texto, str):
            return ""
        return unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('ASCII').strip().lower()

    def carregar_ids_monstros(caminho_arquivo):
        df = pd.read_excel(caminho_arquivo)
        dicionario = {}

        for _, row in df.iterrows():
            nome = normalizar(str(row['nome_monstro']))
            id_monstro = row['id']
            dicionario[nome] = id_monstro

        return dicionario
    

    def buscar_ids_para_monstros(lista_nomes, dict_ids):
        resultado = []

        for nome in lista_nomes:
            nome_limpo = normalizar(nome)
            id_encontrado = dict_ids.get(nome_limpo)
            resultado.append({
                "nome": nome,
                "id": id_encontrado
            })

        return resultado
    dicionario_ids = carregar_ids_monstros(dicionario_ids)
    
    dados_com_ids = buscar_ids_para_monstros(monstros, dicionario_ids)

    data = list(zip(dados_com_ids,base_exp,job_exp,quantidade,mapa,xp_ajustada,nivel_jogador,nivel_monstro))

    return render_template(
        'spots.html',
        dados = data,
        links=carregar_links(),
        streamers = stream_cache
)


@app.route('/clips', methods=['GET', 'POST'])
def clips_page():
    if request.method == 'POST':
        url = request.form.get('url')
        if url:
            url_existente = Link.query.filter_by(url=url).first()
            if url_existente:
                flash("Esse link já foi enviado anteriormente.","erro")
            else:
                novo_link = Link(url=url)
                db.session.add(novo_link)
                db.session.commit()
                flash("Obrigado por compartilhar o clipe, em breve ele será aprovado!","sucesso")
            return redirect('/clips')
    
    links_clipes = Link.query.filter_by(aprovado=True).order_by(Link.data.desc()).all()

    clips_embed = []
    def extrair_slug_twitch(url):
        try:
            if not url:
                return None

            if not url.startswith("http"):
                url = "https://" + url

            parsed = urlparse(url)
            partes = parsed.path.strip("/").split("/")

            if parsed.netloc == "clips.twitch.tv" and len(partes) == 1:
                return partes[0]

            if "clip" in partes:
                idx = partes.index("clip")
                if idx + 1 < len(partes):
                    return partes[idx + 1]

        except Exception as e:
            print(f"Erro ao extrair slug: {e}")
        
        return None

    for link in links_clipes:
        slug = extrair_slug_twitch(link.url)
        print(f"[DEBUG] URL: {link.url} → Slug extraído: {slug}")
        if slug:
            clips_embed.append({
                'slug': slug,
                'id': link.id
            })
        else:
            print(f"[!] URL inválida ignorada: {link.url}")

    print(clips_embed)
    return render_template(
        'clips.html',
        links_clipes=clips_embed,
        links=carregar_links(),
        streamers=stream_cache
    )

@app.route('/admin', methods=['GET', 'POST'])
def admin_page():
    if request.method == 'POST':
        link_id = request.form.get('link_id')
        action = request.form.get('action')

        link = Link.query.get(link_id)
        if not link:
            flash("Link não encontrado.", "erro")
            return redirect('/admin')

        if action == 'aprovar':
            aprovado_val = request.form.get('aprovado')
            link.aprovado = True if aprovado_val == 'true' else False
            db.session.commit()
            flash("Status atualizado com sucesso.", "sucesso")
        elif action == 'deletar':
            db.session.delete(link)
            db.session.commit()
            flash("Link deletado.", "sucesso")
        return redirect('/admin')

    links = Link.query.order_by(Link.data.desc()).all()
    return render_template('admin.html', links=links)

@app.route('/ads.txt')
def render_ads():
 
    return send_from_directory(directory=os.path.abspath("."), path="ads.txt", mimetype='text/plain')


if __name__ == '__main__':
 with app.app_context():
        db.create_all()
 app.run(debug=True)

